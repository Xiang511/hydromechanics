import openpyxl
import requests
import json
import numpy
# 填入api key
API_KEY = ""
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

player_tags = [
 "%232008Q2CVR",
"%23200VYVV8C",
]
wb = openpyxl.Workbook()
ws = wb.create_sheet("Player Data")

# title

ws["A1"] = "Tag"
ws["B1"] = "Name"
ws["C1"] = "天梯排名(old)"
ws["D1"] = "BattleCount"
ws["E1"] = "ThreeCrownWins"
ws["F1"] = "TotalDonations"
ws["G1"] = "tournamentBattleCount"
ws["H1"] = "Winrate"

# write player data
count=1
for player_tag in player_tags:
    response = requests.get(f"https://api.clashroyale.com/v1/players/{player_tag}", headers=headers)
    player_data = response.json()
    x = player_data["wins"]
    y =player_data["losses"]
    z=x/(x+y)
    winrate='{:%}'.format(z)
    tr = json.dumps(player_data.get("leagueStatistics"))
    if tr  != "null":
        if 'bestSeason' in tr:
            bestSeason = eval(tr)["bestSeason"]
            if 'rank' in bestSeason:
                rank = bestSeason["rank"]
            else:
                rank="10000"
    else:
        rank="10000"
    
    count+=1
    print(count)
    
    ws.append([
    player_data["tag"],
    player_data["name"],
    rank,
    player_data["battleCount"],
    player_data["threeCrownWins"],
    player_data["totalDonations"],
    player_data["tournamentBattleCount"],
    winrate
    # player_data["badges"][1]["progress"],
    ])


wb.save("player_data.xlsx")
