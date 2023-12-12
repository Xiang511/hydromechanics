import json
import requests
import openpyxl
import datetime

# Set the API key and headers
API_KEY = ""
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

# Make the request
response = requests.get(
    "https://api.clashroyale.com/v1/locations/57000228/pathoflegend/players",
    headers=headers,
)

# Load the existing Excel workbook
wb = openpyxl.load_workbook("clash_royale_players.xlsx")

# Create a new worksheet
ws = wb.create_sheet("Path of Legends Players")

# Write the header row
# ws.cell(row=1, column=1).value = "Rank"
ws.cell(row=1, column=1).value = "Name"
ws.cell(row=1, column=2).value = "Rating"

now = datetime.datetime.now()
now_str = now.strftime("%Y-%m-%d %H:%M:%S")
ws["E1"] = "執行時間"
ws["E2"] = now_str

# Write the player data
row_number = 2
for player in response.json()["items"]:
    # ws.cell(row=row_number, column=1).value = player["rank"]
    ws.cell(row=row_number, column=1).value = player["name"]
    ws.cell(row=row_number, column=2).value = player["eloRating"]
    row_number += 1

# Save the existing Excel workbook
wb.save("clash_royale_players.xlsx")
print("執行完成")