import openpyxl
import requests
import json
import re
import time
start_time = time.time()
API_KEY = ""
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

player_tags = [
"%232008Q2CVR",
"%23200VYVV8C",
 ] 

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "遊戲ID"
ws["B1"] = "最佳賽季"
ws["C1"] = "Rating"
ws["D1"] = "上一賽季"
ws["E1"] = "Rating"


def clean_result(result):
    if result is None:
        return ""
    return result[result.rfind(" ") + 1:].split("}")[0]



def keep_numbers(string):
  
  filtered_string = filter(str.isdigit, string)
  return "".join(filtered_string)

def get_numbers_by_position_and_delete(variable, start_position, end_position):

  numbers = str(variable)[start_position:end_position]
  filtered_numbers = filter(str.isdigit, numbers)
  return "".join(filtered_numbers)

x=0

for player_tag in player_tags:
    response = requests.get(f"https://api.clashroyale.com/v1/players/{player_tag}", headers=headers)
    
    player_data = response.json()
    best = json.dumps(player_data.get("bestPathOfLegendSeasonResult"))
    last = json.dumps(player_data.get("lastPathOfLegendSeasonResult"))

    best2 = clean_result(best)
    last2= clean_result(last)

    best3 = get_numbers_by_position_and_delete(keep_numbers(best), 2, 6)
    last3 = get_numbers_by_position_and_delete(keep_numbers(last), 2, 6)

    y=player_data["tag"]

    
    

    ws.append([
    player_data["name"],
    best2,best3,
    last2,last3,y])

    x+=1
    print(x)
wb.save("TCRS.xlsx")

end_time = time.time()
print(f"執行時間：{end_time - start_time}")