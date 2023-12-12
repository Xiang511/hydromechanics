import openpyxl
import requests
import json
import re
import time
start_time = time.time()
API_KEY = ""
# 填入api key
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

player_tags = ["%2322R920J00", "%23V99082R9C"]

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "遊戲ID"
ws["B1"] = "最佳賽季"
ws["C1"] = "上一賽季"


def clean_result(result):
    if result is None:
        return ""
    return result[result.rfind(" ") + 1:].split("}")[0]

for player_tag in player_tags:
    response = requests.get(f"https://api.clashroyale.com/v1/players/{player_tag}", headers=headers)
    
    player_data = response.json()
    best = json.dumps(player_data.get("bestPathOfLegendSeasonResult"))
    last = json.dumps(player_data.get("lastPathOfLegendSeasonResult"))

    best2 = clean_result(best)
    last2= clean_result(last)

    ws.append([
    player_data["name"],
    best2,
    last2])
wb.save("clanMember.xlsx")

end_time = time.time()
print(f"執行時間：{end_time - start_time}")